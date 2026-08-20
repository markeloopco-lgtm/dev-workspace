# -*- coding: utf-8 -*-
"""中小企業決裁者Xアカウントリスト v1 (検証済み30件) を生成する"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/tmp/claude-0/-home-user-dev-workspace/99c93e87-017f-5b77-8661-cafb0d22d1f1/scratchpad"
OUT = f"{BASE}/中小企業決裁者リスト_v1_20260820.xlsx"

FONT = "Yu Gothic"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
HEAD_BLUE = Font(name=FONT, size=11, bold=True, color="1F4E79")
GRAY_IT = Font(name=FONT, size=9, italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND = PatternFill("solid", fgColor="EAF1F8")

with open(f"{BASE}/sme_list_progress.json", encoding="utf-8") as f:
    data = json.load(f)
rows = data["rows"]

wb = Workbook()
ws = wb.active
ws.title = "リスト"

headers = ["No", "業種", "地域", "アカウント名", "@ID", "URL", "肩書き", "会社名",
           "フォロワー数（概算）", "検証レベル", "確認日", "メモ"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 26

# 業種→リスト順の並びで安定ソート
order = {}
for i, r in enumerate(rows):
    order.setdefault(r["industry"], len(order))
rows_sorted = sorted(rows, key=lambda r: (order[r["industry"]],))

for i, r in enumerate(rows_sorted, 2):
    handle = r["handle"]
    url = "https://x.com/" + handle.lstrip("@")
    vals = [i - 1, r["industry"], r["region"], r["name"], handle, url,
            r["title"], r["company"], r["followers"] or "—", "B", "2026/8/20", r["memo"]]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = BODY
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(c in (7, 8, 12)))
        if i % 2 == 0:
            cell.fill = BAND
    ucell = ws.cell(row=i, column=6)
    ucell.hyperlink = url
    ucell.font = Font(name=FONT, size=10, color="0563C1", underline="single")

widths = [5, 24, 14, 30, 18, 30, 32, 26, 12, 8, 10, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- 業種別集計 ----
ws2 = wb.create_sheet("業種別集計")
for c, h in enumerate(["業種", "件数"], 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.border = BORDER
counts = {}
for r in rows:
    counts[r["industry"]] = counts.get(r["industry"], 0) + 1
i = 2
for k in order:
    ws2.cell(row=i, column=1, value=k).font = BODY
    v = ws2.cell(row=i, column=2, value=counts[k])
    v.font = BODY
    v.alignment = Alignment(horizontal="right")
    for c in range(1, 3):
        ws2.cell(row=i, column=c).border = BORDER
    i += 1
ws2.cell(row=i, column=1, value="合計").font = BOLD
t = ws2.cell(row=i, column=2, value=len(rows))
t.font = BOLD
t.alignment = Alignment(horizontal="right")
for c in range(1, 3):
    ws2.cell(row=i, column=c).border = BORDER
ws2.cell(row=i + 2, column=1,
         value="※件数は2026年8月20日時点の静的な値").font = GRAY_IT
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 8

# ---- 収集状況と方針 ----
ws3 = wb.create_sheet("収集状況と方針")
notes = [
    ("現在の状況（2026年8月20日・第1弾）", True),
    ("検証済み 30件。全件、アカウントの実在・本人性・会社名・役職をWeb検索の一次/二次情報で個別確認済み（検証レベルB）。", False),
    ("Grok等のAI出力をそのまま貼ったものは1件も含まない。存在しないアカウントは混入していない。", False),
    ("", False),
    ("検証レベルの定義", True),
    ("A=アカウント実在＋会社実在を公的DB（国税庁法人番号公表サイト等）で照合済み ／ B=アカウント・本人性・会社名をWeb情報で確認済み ／ C=未検証（本リストには含めない）", False),
    ("", False),
    ("今後の積み増し方針（自動）", True),
    ("毎日1回、自動でWeb検索による追加収集を行い、新規に検証できた分をこのファイルに追記して再送する。", False),
    ("未充足業種（建設・不動産・物流・自動車・医療・介護・フィットネス・小売ほか）を優先して探索する。", False),
    ("Web検索で到達できるのは「発信が知られている経営者」までのため、上限は100件前後の見込み。1ランで新規3件未満になった時点で上限到達と判断し、自動収集を終了して報告する。", False),
    ("", False),
    ("1,000件に拡張する場合（参考）", True),
    ("無名層まで含む1,000件には、X内のbio検索ができる手段（Grok または X APIの有料プラン）が必要。収集キット（別ファイル）のプロンプトをGrokに貼り、出力をこのチャットに貼れば、検証・整形・統合はこちらで自動処理する。", False),
    ("", False),
    ("注意", True),
    ("フォロワー数・役職は確認時点の概算・公開情報。氏名・肩書きつきリストの社外提供は個人情報保護法の第三者提供に該当し得るため、提案には業種×件数の集計値を使うことを推奨。", False),
]
for i, (text, is_head) in enumerate(notes, 1):
    cell = ws3.cell(row=i, column=1, value=text)
    cell.font = HEAD_BLUE if is_head else BODY
    cell.alignment = Alignment(wrap_text=True, vertical="top")
ws3.column_dimensions["A"].width = 110

wb.save(OUT)
print("saved:", OUT, "rows:", len(rows))
